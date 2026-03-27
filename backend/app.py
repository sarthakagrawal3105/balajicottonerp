import streamlit as st
import openpyxl
import random
import io
import pandas as pd
import datetime
import zipfile
from openpyxl.drawing.image import Image as OpenpyxlImage

st.set_page_config(page_title="CCI Multi-Bill Generator", layout="wide", page_icon="📝")

st.title("📝 CCI Multiple Bale Bills Generator")
st.markdown("Generate multiple exact-match Excel sheets into a **single .xlsx file** with your **LOGO fully preserved!**")

# 1. Global Settings
st.subheader("1. Global Bounds Settings")
col1, col2, col3 = st.columns(3)
with col1:
    min_wt = st.number_input("Minimum Weight (KG)", value=160.0)
with col2:
    max_wt = st.number_input("Maximum Weight (KG)", value=178.0)
with col3:
    tare_weight = st.number_input("Tare Weight", value=1.10, step=0.01)

uploaded_file = st.file_uploader("Upload Excel Template", type=["xlsx"])

# 2. Bills Data Entry
st.subheader("2. Bills Data Entry")
n_bills = st.number_input("How many sheets do you want to create in the file?", min_value=1, value=1, step=1)

today_str = datetime.date.today().strftime("%d-%m-%Y")

# Generate default configuration if size changes
if "n_bills_state" not in st.session_state or st.session_state.n_bills_state != n_bills:
    default_data = {
        "Sheet Name": [f"SPL {i+1}" for i in range(int(n_bills))],
        "Date": [today_str] * int(n_bills),
        "PR Number": [f"{i*100 + 1}-{(i+1)*100}" for i in range(int(n_bills))],
        "Lot Number": ["4001"] * int(n_bills),
        "Gross Weight (QTL)": [170.05] * int(n_bills),
        "Number of Bales": [100] * int(n_bills)
    }
    
    st.session_state.df_input = pd.DataFrame(default_data)
    st.session_state.n_bills_state = n_bills

st.markdown("Fill out the specific details for each bill below. Everything will be merged into **ONE** Excel file!")
edited_df = st.data_editor(st.session_state.df_input, num_rows="dynamic", use_container_width=True)

def generate_weights(num, target_sum, min_w, max_w):
    bales = [random.randint(min_w, max_w) for _ in range(num)]
    current_sum = sum(bales)
    while current_sum != target_sum:
        i = random.randint(0, num - 1)
        if current_sum < target_sum:
            if bales[i] < max_w:
                bales[i] += 1
                current_sum += 1
        elif current_sum > target_sum:
            if bales[i] > min_w:
                bales[i] -= 1
                current_sum -= 1
    random.shuffle(bales)
    return bales

if st.button("🚀 Generate Single Combined Excel File", type="primary"):
    if not uploaded_file:
        st.error("Please upload the Excel template first.")
    elif min_wt > max_wt:
        st.error("Minimum weight cannot be greater than maximum weight.")
    else:
        try:
            error_found = False
            
            with st.spinner("Generating weights and building Excel sheets..."):
                # Extract image bytes directly from ZIP to perfectly preserve logos
                uploaded_file.seek(0)
                file_bytes = uploaded_file.read()
                
                image_bytes_list = []
                with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
                    for item in archive.namelist():
                        if item.startswith('xl/media/'):
                            image_bytes_list.append(archive.read(item))
                
                # Load workbook
                uploaded_file.seek(0)
                wb = openpyxl.load_workbook(uploaded_file)
                template_sheet = wb.active # Use the first active sheet as template
                old_sheets = wb.sheetnames # Keep track of ALL original sheets to delete later
                
                # Fetch original image anchors
                anchors = [img.anchor for img in template_sheet._images]
                
                for index, row in edited_df.iterrows():
                    sheet_name = str(row["Sheet Name"])
                    bill_date = str(row["Date"])
                    pr_no = str(row["PR Number"])
                    lot_no = str(row["Lot Number"])
                    gross_weight = float(row["Gross Weight (QTL)"])
                    num_bales = int(row["Number of Bales"])
                    
                    target_kgs = int(round(gross_weight * 100))
                    if target_kgs < min_wt * num_bales or target_kgs > max_wt * num_bales:
                        st.error(f"Row {index+1} ({sheet_name}): Target sum ({target_kgs} KG) is not possible within the {min_wt}-{max_wt} KG bounds for {num_bales} bales!")
                        error_found = True
                        break
                    
                    bales = generate_weights(num_bales, target_kgs, int(min_wt), int(max_wt))
                    
                    # Copy template for this specific bill
                    target_sheet = wb.copy_worksheet(template_sheet)
                    target_sheet.title = sheet_name
                    
                    # Apply raw images back manually (openpyxl deletes them on copy)
                    for i_img, img_data in enumerate(image_bytes_list):
                        if i_img < len(anchors):
                            fresh_img = OpenpyxlImage(io.BytesIO(img_data))
                            target_sheet.add_image(fresh_img, anchors[i_img])
                    
                    # Update Static Header fields
                    target_sheet["C8"] = pr_no
                    target_sheet["C9"] = bill_date
                    target_sheet["K10"] = lot_no
                    target_sheet["G11"] = num_bales
                    
                    # Fill Bales 1-60
                    for r in range(10):
                        for c in range(6):
                            bale_idx = c * 10 + r
                            col_letter = chr(ord('B') + c*2)
                            if bale_idx < num_bales:
                                target_sheet[f"{col_letter}{14 + r}"] = bales[bale_idx]
                            else:
                                target_sheet[f"{col_letter}{14 + r}"] = ""
                    
                    # Block totals 1-60 (Formulas)
                    for c in range(6):
                        col_letter = chr(ord('B') + c*2)
                        target_sheet[f"{col_letter}24"] = f"=SUM({col_letter}14:{col_letter}23)" if c*10 < num_bales else ""
                    
                    # Fill Bales 61-100
                    for r in range(10):
                        for c in range(6, 10):
                            bale_idx = c * 10 + r
                            col_letter = chr(ord('B') + (c-6)*2)
                            if bale_idx < num_bales:
                                target_sheet[f"{col_letter}{28 + r}"] = bales[bale_idx]
                            else:
                                target_sheet[f"{col_letter}{28 + r}"] = ""
                                
                    # Block totals 61-100 (Formulas)
                    for c in range(6, 10):
                        col_letter = chr(ord('B') + (c-6)*2)
                        target_sheet[f"{col_letter}38"] = f"=SUM({col_letter}28:{col_letter}37)" if c*10 < num_bales else ""
                    
                    # Vertical Summary table (Formulas)
                    sources = ['B24', 'D24', 'F24', 'H24', 'J24', 'L24', 'B38', 'D38', 'F38', 'H38']
                    for b in range(10):
                        if b*10 < num_bales:
                            target_sheet[f"J{28 + b}"] = f"={sources[b]}"
                        else:
                            target_sheet[f"J{28 + b}"] = ""
                    
                    # Grand total (Formula)
                    target_sheet["J38"] = "=SUM(J28:J37)"
                    
                    # Footers (Fully dynamic based on J38 grand total)
                    target_sheet["K40"] = "=J38/100"
                    target_sheet["K41"] = tare_weight
                    target_sheet["K42"] = "=K40-K41"
            
                if not error_found:
                    # STRICTLY Delete ALL old sheets from the workbook, 
                    # so only the freshly generated ones remain!
                    for old_sh in old_sheets:
                        if old_sh in wb.sheetnames:
                            del wb[old_sh]
                    
                    # Export single combined Excel memory buffer
                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)
                    
                    st.success(f"✅ Generated {len(edited_df)} bills in a SINGLE Excel file successfully!")
                    
                    st.download_button(
                        label="📥 Download Finished Combined Excel File (.xlsx)",
                        data=output,
                        file_name="cci_combined_bills.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        except Exception as e:
            st.error(f"Error modifying Excel: {e}")
